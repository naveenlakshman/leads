# app.py
from datetime import datetime, date
from flask import Flask, render_template, request, redirect, url_for, flash, abort, send_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_migrate import Migrate
from config import Config
from models import db, User, Lead, FollowUp, Activity
from utils.auth import admin_required
from utils.helpers import parse_date, utc_to_ist, log_activity
from utils.lead_score import compute_lead_score
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

def create_app():
    app = Flask(__name__)
    app.config.from_object(Config)

    # Initialize Flask-Migrate
    migrate = Migrate(app, db)

    # predefined dropdown options used across lead forms
    GENDER_OPTIONS = ["Male", "Female", "Other"]
    EDUCATION_OPTIONS = ["Below 10th", "SSLC", "PUC", "Degree", "Working", "Job seeker"]
    STREAM_OPTIONS = ["Commerce", "Science", "Arts"]
    CAREER_GOAL_OPTIONS = ["Job", "Internship", "Skills", "Business"]
    LEAD_SOURCE_OPTIONS = ["Walk-in", "Instagram", "Referral","Google Call","Poster" , "Other",]
    TIMEFRAME_OPTIONS = ["Immediately", "1 week", "1 month", "Exploring"]
    DECISION_MAKER_OPTIONS = ["Self", "Parents", "Friends"]

    FOLLOWUP_METHODS = ["Call", "WhatsApp", "Email", "In-person", "Other"]
    FOLLOWUP_OUTCOMES = ["Interested", "Call back", "Not interested", "No answer", "Other"]

    # Ensure instance folder exists
    import os
    os.makedirs(app.config["INSTANCE_DIR"], exist_ok=True)

    # Init DB
    db.init_app(app)

    # Login manager
    login_manager = LoginManager()
    login_manager.login_view = "login"
    login_manager.init_app(app)

    @login_manager.user_loader
    def load_user(user_id):
        return User.query.get(int(user_id))

    # Jinja filter for IST display
    @app.template_filter("ist")
    def ist_filter(dt):
        x = utc_to_ist(dt)
        if not x:
            return ""
        return x.strftime("%d-%b-%Y %I:%M %p")

    # Create tables + default users once
    with app.app_context():
        db.create_all()
        seed_default_users()

    # -----------------------
    # AUTH ROUTES
    # -----------------------
    @app.route("/")
    def index():
        if current_user.is_authenticated:
            return redirect(url_for("dashboard"))
        return redirect(url_for("login"))

    @app.route("/login", methods=["GET", "POST"])
    def login():
        if current_user.is_authenticated:
            return redirect(url_for("dashboard"))

        if request.method == "POST":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "").strip()

            user = User.query.filter_by(username=username).first()
            if not user or not user.check_password(password):
                flash("Invalid username or password", "danger")
                return render_template("login.html")

            if not user.is_active:
                flash("Your account is disabled. Contact admin.", "danger")
                return render_template("login.html")

            login_user(user)
            return redirect(url_for("dashboard"))

        return render_template("login.html")

    @app.route("/logout")
    @login_required
    def logout():
        logout_user()
        return redirect(url_for("login"))

    # -----------------------
    # DASHBOARD
    # -----------------------
    @app.route("/dashboard")
    @login_required
    def dashboard():
        today = date.today()

        # Counselors see only their leads; admins see all by default
        query_filter = (Lead.assigned_to_id == current_user.id) if current_user.role == "counselor" else True

        new_leads_today = Lead.query.filter(
            db.func.date(Lead.created_at) == str(today),
            Lead.is_deleted == False,
            query_filter
        ).count()

        total_leads_this_month = Lead.query.filter(
            db.extract("month", Lead.created_at) == today.month,
            db.extract("year", Lead.created_at) == today.year,
            Lead.is_deleted == False,
            query_filter
        ).count()

        followups_due = Lead.query.filter(
            Lead.status == "active",
            Lead.next_followup_date.isnot(None),
            Lead.next_followup_date <= today,
            Lead.is_deleted == False,
            query_filter
        ).order_by(Lead.next_followup_date.asc()).all()

        hot_leads = Lead.query.filter(
            Lead.status == "active",
            Lead.lead_score >= 60,
            Lead.is_deleted == False,
            query_filter
        ).order_by(Lead.lead_score.desc()).limit(10).all()

        converted_this_month = Lead.query.filter(
            Lead.status == "converted",
            db.extract("month", Lead.updated_at) == today.month,
            db.extract("year", Lead.updated_at) == today.year,
            Lead.is_deleted == False,
            query_filter
        ).count()

        total_active = Lead.query.filter(Lead.status == "active", Lead.is_deleted == False, query_filter).count()
        active_this_month = Lead.query.filter(
            Lead.status == "active",
            db.extract("month", Lead.created_at) == today.month,
            db.extract("year", Lead.created_at) == today.year,
            Lead.is_deleted == False,
            query_filter
        ).count()
        
        # Enhanced metrics
        total_leads = Lead.query.filter(Lead.is_deleted == False, query_filter).count()
        converted_total = Lead.query.filter(Lead.status == "converted", Lead.is_deleted == False, query_filter).count()
        lost_total = Lead.query.filter(Lead.status == "lost", Lead.is_deleted == False, query_filter).count()
        lost_this_month = Lead.query.filter(
            Lead.status == "lost",
            db.extract("month", Lead.updated_at) == today.month,
            db.extract("year", Lead.updated_at) == today.year,
            Lead.is_deleted == False,
            query_filter
        ).count()
        
        # Conversion rate
        conversion_rate = round((converted_total / total_leads * 100), 1) if total_leads > 0 else 0
        
        # Stage breakdown
        stage_breakdown = db.session.query(
            Lead.stage,
            db.func.count(Lead.id)
        ).filter(Lead.is_deleted == False, query_filter, Lead.status == "active").group_by(Lead.stage).all()
        
        # High-risk leads (old last contact)
        high_risk_leads = Lead.query.filter(
            Lead.status == "active",
            Lead.is_deleted == False,
            query_filter
        ).order_by(Lead.last_contact_date.desc()).limit(5).all()
        
        # For admins: team performance summary
        team_stats = None
        if current_user.role == "admin":
            team_stats = []
            counselors = User.query.filter(User.role == "counselor", User.is_active == True).all()
            
            for counselor in counselors:
                c_total = Lead.query.filter(Lead.assigned_to_id == counselor.id, Lead.is_deleted == False).count()
                c_converted = Lead.query.filter(Lead.assigned_to_id == counselor.id, Lead.status == "converted", Lead.is_deleted == False).count()
                c_rate = round((c_converted / c_total * 100), 1) if c_total > 0 else 0
                
                if c_total > 0:
                    team_stats.append({
                        'name': counselor.full_name or counselor.username,
                        'total': c_total,
                        'converted': c_converted,
                        'rate': c_rate
                    })
            
            # Sort by conversion rate
            team_stats.sort(key=lambda x: x['rate'], reverse=True)

        return render_template(
            "dashboard.html",
            new_leads_today=new_leads_today,
            followups_due=followups_due[:10],
            followups_due_count=len(followups_due),
            hot_leads=hot_leads,
            converted_this_month=converted_this_month,
            total_active=total_active,
            active_this_month=active_this_month,
            total_leads=total_leads,
            total_leads_this_month=total_leads_this_month,
            converted_total=converted_total,
            lost_total=lost_total,
            lost_this_month=lost_this_month,
            conversion_rate=conversion_rate,
            stage_breakdown=stage_breakdown,
            high_risk_leads=high_risk_leads,
            team_stats=team_stats,
            is_admin=current_user.role == "admin",
            now=today
        )

    # -----------------------
    # LEADS
    # -----------------------
    @app.route("/leads")
    @login_required
    def leads_list():
        q = request.args.get("q", "").strip()
        stage = request.args.get("stage", "").strip()
        source = request.args.get("source", "").strip()
        user_id = request.args.get("user_id", "").strip()
        today = date.today()

        base_query = Lead.query.filter(Lead.is_deleted == False)

        # Counselors see only their leads; admins see all by default or filter by user_id
        if current_user.role == "counselor":
            base_query = base_query.filter(Lead.assigned_to_id == current_user.id)
            all_users = [current_user]  # Counselors only see themselves
        else:
            # Admin: show all by default, or filter by specific user if user_id is provided
            if user_id:
                try:
                    base_query = base_query.filter(Lead.assigned_to_id == int(user_id))
                except (ValueError, TypeError):
                    pass  # Invalid user_id, show all
            
            # Get all users (including inactive) for the dropdown
            all_users = User.query.order_by(User.full_name).all()

        metrics = {
            "total_overall": base_query.count(),
            "total_this_month": base_query.filter(
                db.extract("month", Lead.created_at) == today.month,
                db.extract("year", Lead.created_at) == today.year
            ).count(),
            "active_overall": base_query.filter(Lead.status == "active").count(),
            "active_this_month": base_query.filter(
                Lead.status == "active",
                db.extract("month", Lead.created_at) == today.month,
                db.extract("year", Lead.created_at) == today.year
            ).count(),
            "converted_overall": base_query.filter(Lead.status == "converted").count(),
            "converted_this_month": base_query.filter(
                Lead.status == "converted",
                db.extract("month", Lead.updated_at) == today.month,
                db.extract("year", Lead.updated_at) == today.year
            ).count(),
            "lost_overall": base_query.filter(Lead.status == "lost").count(),
            "lost_this_month": base_query.filter(
                Lead.status == "lost",
                db.extract("month", Lead.updated_at) == today.month,
                db.extract("year", Lead.updated_at) == today.year
            ).count(),
        }

        query = base_query
        if q:
            like = f"%{q}%"
            query = query.filter(
                db.or_(
                    Lead.name.ilike(like),
                    Lead.phone.ilike(like),
                    Lead.whatsapp.ilike(like)
                )
            )
        if stage:
            query = query.filter(Lead.stage == stage)
        if source:
            query = query.filter(Lead.lead_source == source)

        leads = query.order_by(Lead.updated_at.desc()).all()

        # Dropdown values
        stages = ["New Lead", "Contacted", "Interested", "Counseling Done", "Follow-up", "Converted", "Lost"]
        sources = LEAD_SOURCE_OPTIONS

        return render_template("leads.html", leads=leads, q=q, stage=stage, source=source, stages=stages, sources=sources, 
                             is_admin=(current_user.role == "admin"), all_users=all_users, 
                             selected_user_id=user_id if user_id else None, metrics=metrics)

    @app.route("/admin/deleted-leads")
    @login_required
    @admin_required
    def deleted_leads():
        deleted = Lead.query.filter(Lead.is_deleted == True).order_by(Lead.updated_at.desc()).all()
        return render_template("deleted_leads.html", leads=deleted)

    @app.route("/leads/new", methods=["GET", "POST"])
    @login_required
    def lead_create():
        if request.method == "POST":
            lead = Lead(
                name=request.form.get("name", "").strip(),
                phone=request.form.get("phone", "").strip(),
                whatsapp=request.form.get("whatsapp", "").strip() or None,
                gender=request.form.get("gender", "").strip() or None,
                age=int(request.form.get("age")) if request.form.get("age") else None,
                education_status=request.form.get("education_status", "").strip() or None,
                stream=request.form.get("stream", "").strip() or None,
                institute_name=request.form.get("institute_name", "").strip() or None,
                career_goal=request.form.get("career_goal", "").strip() or None,
                interested_courses=request.form.get("interested_courses", "").strip() or None,
                lead_source=request.form.get("lead_source", "").strip() or None,
                decision_maker=request.form.get("decision_maker", "Self").strip() or "Self",
                start_timeframe=request.form.get("start_timeframe", "").strip() or None,
                stage=request.form.get("stage", "New Lead").strip() or "New Lead",
                notes=request.form.get("notes", "").strip() or None,
            )

            lead.last_contact_date = parse_date(request.form.get("last_contact_date"))
            lead.next_followup_date = parse_date(request.form.get("next_followup_date"))

            lead.lead_score = compute_lead_score(lead.lead_source, lead.start_timeframe, lead.education_status, lead.career_goal)

            # assign to current user by default
            lead.assigned_to_id = current_user.id
            
            # set status based on stage
            if lead.stage == "Converted":
                lead.status = "converted"
                lead.next_followup_date = None
            elif lead.stage == "Lost":
                lead.status = "lost"
                lead.next_followup_date = None
            else:
                lead.status = "active"

            if not lead.name or not lead.phone:
                flash("Name and Phone are required.", "danger")
                return render_template(
                    "lead_form.html",
                    lead=None,
                    mode="create",
                    genders=GENDER_OPTIONS,
                    educations=EDUCATION_OPTIONS,
                    streams=STREAM_OPTIONS,
                    career_goals=CAREER_GOAL_OPTIONS,
                    lead_sources=LEAD_SOURCE_OPTIONS,
                    decision_makers=DECISION_MAKER_OPTIONS,
                    timeframes=TIMEFRAME_OPTIONS,
                )

            db.session.add(lead)
            db.session.commit()
            
            # Log activity
            log_activity(
                user_id=current_user.id,
                lead_id=lead.id,
                action_type="lead_created",
                description=f"Lead created: {lead.name} ({lead.phone}) - Stage: {lead.stage}, Source: {lead.lead_source}"
            )
            
            flash("Lead created successfully.", "success")
            return redirect(url_for("leads_list"))

        return render_template(
            "lead_form.html",
            lead=None,
            mode="create",
            genders=GENDER_OPTIONS,
            educations=EDUCATION_OPTIONS,
            streams=STREAM_OPTIONS,
            career_goals=CAREER_GOAL_OPTIONS,
            lead_sources=LEAD_SOURCE_OPTIONS,
            decision_makers=DECISION_MAKER_OPTIONS,
            timeframes=TIMEFRAME_OPTIONS,
        )

    @app.route("/leads/<int:lead_id>")
    @login_required
    def lead_detail(lead_id):
        lead = Lead.query.get_or_404(lead_id)
        all_users = User.query.filter(User.is_active == True).order_by(User.full_name).all()
        return render_template(
            "lead_detail.html",
            lead=lead,
            all_users=all_users,
            methods=FOLLOWUP_METHODS,
            outcomes=FOLLOWUP_OUTCOMES,
        )

    @app.route("/leads/<int:lead_id>/edit", methods=["GET", "POST"])
    @login_required
    def lead_edit(lead_id):
        lead = Lead.query.get_or_404(lead_id)

        if request.method == "POST":
            lead.name = request.form.get("name", "").strip()
            lead.phone = request.form.get("phone", "").strip()
            lead.whatsapp = request.form.get("whatsapp", "").strip() or None
            lead.gender = request.form.get("gender", "").strip() or None
            lead.age = int(request.form.get("age")) if request.form.get("age") else None

            lead.education_status = request.form.get("education_status", "").strip() or None
            lead.stream = request.form.get("stream", "").strip() or None
            lead.institute_name = request.form.get("institute_name", "").strip() or None

            lead.career_goal = request.form.get("career_goal", "").strip() or None
            lead.interested_courses = request.form.get("interested_courses", "").strip() or None
            lead.lead_source = request.form.get("lead_source", "").strip() or None
            lead.decision_maker = request.form.get("decision_maker", "Self").strip() or "Self"
            lead.start_timeframe = request.form.get("start_timeframe", "").strip() or None

            lead.stage = request.form.get("stage", lead.stage).strip() or lead.stage
            lead.notes = request.form.get("notes", "").strip() or None

            lead.last_contact_date = parse_date(request.form.get("last_contact_date"))
            lead.next_followup_date = parse_date(request.form.get("next_followup_date"))

            # recompute score
            lead.lead_score = compute_lead_score(lead.lead_source, lead.start_timeframe, lead.education_status, lead.career_goal)

            # status auto-sync
            if lead.stage == "Converted":
                lead.status = "converted"
                lead.next_followup_date = None
            elif lead.stage == "Lost":
                lead.status = "lost"
                lead.next_followup_date = None
            else:
                lead.status = "active"

            db.session.commit()
            
            # Log activity
            log_activity(
                user_id=current_user.id,
                lead_id=lead.id,
                action_type="lead_edited",
                description=f"Lead updated: {lead.name} - Current Stage: {lead.stage}"
            )
            
            flash("Lead updated.", "success")
            return redirect(url_for("lead_detail", lead_id=lead.id))

        return render_template(
            "lead_form.html",
            lead=lead,
            mode="edit",
            genders=GENDER_OPTIONS,
            educations=EDUCATION_OPTIONS,
            streams=STREAM_OPTIONS,
            career_goals=CAREER_GOAL_OPTIONS,
            lead_sources=LEAD_SOURCE_OPTIONS,
            decision_makers=DECISION_MAKER_OPTIONS,
            timeframes=TIMEFRAME_OPTIONS,
        )

    @app.route("/leads/<int:lead_id>/delete", methods=["POST"])
    @login_required
    def lead_delete(lead_id):
        lead = Lead.query.get_or_404(lead_id)
        lead.is_deleted = True
        db.session.commit()
        
        # Log activity
        log_activity(
            user_id=current_user.id,
            lead_id=lead.id,
            action_type="lead_deleted",
            description=f"Lead deactivated: {lead.name}"
        )
        
        flash("Lead deactivated. (Data is preserved and can be viewed by admins)", "warning")
        return redirect(url_for("leads_list"))

    @app.route("/leads/<int:lead_id>/convert", methods=["POST"])
    @login_required
    def lead_convert(lead_id):
        lead = Lead.query.get_or_404(lead_id)
        lead.stage = "Converted"
        lead.status = "converted"
        lead.next_followup_date = None
        db.session.commit()
        
        # Log activity
        log_activity(
            user_id=current_user.id,
            lead_id=lead.id,
            action_type="lead_converted",
            description=f"Lead marked as Converted: {lead.name}"
        )
        
        flash("Lead marked as Converted.", "success")
        return redirect(url_for("lead_detail", lead_id=lead.id))

    @app.route("/leads/<int:lead_id>/mark_lost", methods=["POST"])
    @login_required
    def lead_mark_lost(lead_id):
        lead = Lead.query.get_or_404(lead_id)
        reason = request.form.get("lost_reason", "").strip() or None
        lead.stage = "Lost"
        lead.status = "lost"
        lead.lost_reason = reason
        lead.next_followup_date = None
        db.session.commit()
        
        # Log activity
        log_activity(
            user_id=current_user.id,
            lead_id=lead.id,
            action_type="lead_lost",
            description=f"Lead marked as Lost: {lead.name} - Reason: {reason or 'Not specified'}"
        )
        
        flash("Lead marked as Lost.", "warning")
        return redirect(url_for("lead_detail", lead_id=lead.id))

    @app.route("/leads/<int:lead_id>/reassign", methods=["POST"])
    @login_required
    def lead_reassign(lead_id):
        lead = Lead.query.get_or_404(lead_id)
        assigned_to_id = request.form.get("assigned_to_id", "").strip() or None
        
        if assigned_to_id:
            # Verify user exists and is active
            user = User.query.get(int(assigned_to_id))
            if not user or not user.is_active:
                flash("Invalid user selected.", "danger")
                return redirect(url_for("lead_detail", lead_id=lead.id))
            lead.assigned_to_id = int(assigned_to_id)
            flash(f"Lead reassigned to {user.full_name or user.username}.", "success")
        else:
            lead.assigned_to_id = None
            flash("Lead unassigned.", "info")
        
        db.session.commit()
        return redirect(url_for("lead_detail", lead_id=lead.id))

    # -----------------------
    # FOLLOWUPS (today list)
    # -----------------------
    @app.route("/followups")
    @login_required
    def followups_today():
        today = date.today()
        
        # Counselors see only their leads; admins can filter by user
        user_filter = request.args.get("user_id", "").strip()
        query_filter = (Lead.assigned_to_id == current_user.id) if current_user.role == "counselor" else True
        
        # If admin selected a specific user, filter by that
        if current_user.role == "admin" and user_filter:
            try:
                query_filter = Lead.assigned_to_id == int(user_filter)
            except (ValueError, TypeError):
                pass
        
        leads = Lead.query.filter(
            Lead.status == "active",
            Lead.next_followup_date.isnot(None),
            Lead.next_followup_date <= today,
            query_filter
        ).order_by(Lead.next_followup_date.asc()).all()

        # Get all active users for admin dropdown
        all_users = User.query.filter(User.is_active == True).order_by(User.full_name).all() if current_user.role == "admin" else []

        return render_template("followups.html", leads=leads, today=today, all_users=all_users, selected_user_id=user_filter, is_admin=(current_user.role == "admin"))

    # Add followup note to a lead
    @app.route("/leads/<int:lead_id>/followups/new", methods=["POST"])
    @login_required
    def followup_add(lead_id):
        lead = Lead.query.get_or_404(lead_id)
        method = request.form.get("method", "").strip() or None
        outcome = request.form.get("outcome", "").strip() or None
        note = request.form.get("note", "").strip() or None
        next_dt = parse_date(request.form.get("next_followup_date"))

        fu = FollowUp(
            lead_id=lead.id,
            user_id=current_user.id,
            method=method,
            outcome=outcome,
            note=note,
            next_followup_date=next_dt
        )
        db.session.add(fu)

        # update lead tracking
        lead.last_contact_date = date.today()
        lead.followup_count = (lead.followup_count or 0) + 1
        lead.next_followup_date = next_dt

        # stage nudge
        if lead.stage == "New Lead":
            lead.stage = "Contacted"

        db.session.commit()
        
        # Log activity
        log_activity(
            user_id=current_user.id,
            lead_id=lead.id,
            action_type="followup_added",
            description=f"Follow-up added for {lead.name} - Method: {method or 'Not specified'}, Outcome: {outcome or 'Not specified'}"
        )
        
        flash("Follow-up saved.", "success")
        return redirect(url_for("lead_detail", lead_id=lead.id))

    # -----------------------
    # PIPELINE
    # -----------------------
    def get_next_stages(current_stage):
        """Return dict of next stage(s) available from current stage."""
        stage_flow = {
            "New Lead": [{"name": "Contacted", "color": "primary"}],
            "Contacted": [{"name": "Interested", "color": "info"}],
            "Interested": [{"name": "Counseling Done", "color": "warning"}],
            "Counseling Done": [{"name": "Follow-up", "color": "secondary"}],
            "Follow-up": [
                {"name": "Converted", "color": "success"},
                {"name": "Lost", "color": "danger"}
            ],
            "Converted": [],  # terminal
            "Lost": []        # terminal
        }
        return stage_flow.get(current_stage, [])

    @app.route("/pipeline")
    @login_required
    def pipeline():
        stages = ["New Lead", "Contacted", "Interested", "Counseling Done", "Follow-up", "Converted", "Lost"]
        
        # Get user_id filter parameter (defaults to None for show all)
        user_id = request.args.get("user_id", "").strip()
        
        # Counselors see only their leads; admins see all by default or filter by user_id
        base_query = Lead.query.filter(Lead.is_deleted == False)
        if current_user.role == "counselor":
            base_query = base_query.filter(Lead.assigned_to_id == current_user.id)
            all_users = [current_user]  # Counselors only see themselves
        else:
            # Admin: show all by default, or filter by specific user if user_id is provided
            if user_id:
                try:
                    user_id = int(user_id)
                    base_query = base_query.filter(Lead.assigned_to_id == user_id)
                except (ValueError, TypeError):
                    pass  # Invalid user_id, show all
            
            # Get all users (including inactive) for the dropdown
            all_users = User.query.order_by(User.full_name).all()
        
        data = {}
        for st in stages:
            data[st] = base_query.filter(Lead.stage == st).order_by(Lead.updated_at.desc()).limit(50).all()
        
        return render_template("pipeline.html", stages=stages, data=data, get_next_stages=get_next_stages, 
                             is_admin=(current_user.role == "admin"), all_users=all_users, 
                             selected_user_id=user_id if user_id else None)

    # quick stage update (buttons)
    @app.route("/leads/<int:lead_id>/stage", methods=["POST"])
    @login_required
    def lead_set_stage(lead_id):
        lead = Lead.query.get_or_404(lead_id)
        st = request.form.get("stage", "").strip()
        if st not in ["New Lead", "Contacted", "Interested", "Counseling Done", "Follow-up", "Converted", "Lost"]:
            abort(400)

        old_stage = lead.stage
        lead.stage = st
        if st == "Converted":
            lead.status = "converted"
            lead.next_followup_date = None
        elif st == "Lost":
            lead.status = "lost"
            lead.next_followup_date = None
        else:
            lead.status = "active"
        db.session.commit()
        
        # Log activity
        log_activity(
            user_id=current_user.id,
            lead_id=lead.id,
            action_type="stage_changed",
            description=f"Lead stage changed: {lead.name} - {old_stage} → {st}",
            field_changed="stage",
            old_value=old_stage,
            new_value=st
        )
        
        return redirect(request.referrer or url_for("pipeline"))

    # -----------------------
    # REPORTS (admin only)
    # -----------------------
    @app.route("/reports", strict_slashes=False)
    @login_required
    @admin_required
    def reports():
        user_id_filter = request.args.get("user_id", "").strip()
        date_from = request.args.get("date_from", "").strip()
        date_to = request.args.get("date_to", "").strip()
        today = date.today()

        # Build base filter
        base_filter = []
        
        # User filter
        if user_id_filter:
            try:
                user_id_filter = int(user_id_filter)
                base_filter.append(Lead.assigned_to_id == user_id_filter)
            except (ValueError, TypeError):
                pass
        
        # Date range filters
        if date_from:
            try:
                from_date = datetime.strptime(date_from, "%Y-%m-%d").date()
                base_filter.append(db.func.date(Lead.created_at) >= from_date)
            except ValueError:
                date_from = ""
        
        if date_to:
            try:
                to_date = datetime.strptime(date_to, "%Y-%m-%d").date()
                base_filter.append(db.func.date(Lead.created_at) <= to_date)
            except ValueError:
                date_to = ""
        
        # Combine all filters with AND
        if base_filter:
            query_filter = db.and_(*base_filter, Lead.is_deleted == False)
        else:
            query_filter = Lead.is_deleted == False

        total_leads = Lead.query.filter(query_filter).count()
        active = Lead.query.filter(Lead.status == "active", query_filter).count()
        converted_total = Lead.query.filter(Lead.status == "converted", query_filter).count()
        lost = Lead.query.filter(Lead.status == "lost", query_filter).count()

        # Overall conversion rate
        conversion_rate = round((converted_total / total_leads * 100), 1) if total_leads > 0 else 0

        # source performance - include all predefined sources with conversion rates
        source_query = db.session.query(
            Lead.lead_source,
            db.func.count(Lead.id),
            db.func.count(db.case((Lead.status == "converted", 1), else_=None))
        ).filter(query_filter, Lead.lead_source.isnot(None)).group_by(Lead.lead_source).all()
        
        # Create a dictionary of source stats (total, converted)
        source_dict = {s: (total, source_converted) for s, total, source_converted in source_query}
        
        # Add all predefined sources with their stats
        source_rows = []
        for source in LEAD_SOURCE_OPTIONS:
            total, source_converted = source_dict.get(source, (0, 0))
            conv_rate = round((source_converted / total * 100), 1) if total > 0 else 0
            source_rows.append((source, total, source_converted, conv_rate))

        # course interest - with conversion rate
        course_rows = db.session.query(
            Lead.interested_courses,
            db.func.count(Lead.id),
            db.func.count(db.case((Lead.status == "converted", 1), else_=None))
        ).filter(query_filter).group_by(Lead.interested_courses).all()
        
        # Calculate conversion rates for courses
        course_rows = [(course, total, converted, round((converted/total*100), 1) if total > 0 else 0) 
                       for course, total, converted in course_rows]

        # Get all users for the dropdown (including inactive)
        all_users = User.query.order_by(User.full_name).all()

        # User Performance Metrics (show all users if no specific filter)
        user_stats = []
        if not user_id_filter:
            # Show summary of all users with leads
            users = all_users
            for user in users:
                user_base_filter = [Lead.assigned_to_id == user.id, Lead.is_deleted == False]
                
                # Apply date filters
                if date_from:
                    try:
                        from_date = datetime.strptime(date_from, "%Y-%m-%d").date()
                        user_base_filter.append(db.func.date(Lead.created_at) >= from_date)
                    except ValueError:
                        pass
                
                if date_to:
                    try:
                        to_date = datetime.strptime(date_to, "%Y-%m-%d").date()
                        user_base_filter.append(db.func.date(Lead.created_at) <= to_date)
                    except ValueError:
                        pass
                
                user_query_filter = db.and_(*user_base_filter) if user_base_filter else True
                user_query = Lead.query.filter(user_query_filter)
                
                user_total = user_query.count()
                user_active = user_query.filter(Lead.status == "active").count()
                user_converted = user_query.filter(Lead.status == "converted").count()
                user_lost = user_query.filter(Lead.status == "lost").count()
                user_conv_rate = round((user_converted / user_total * 100), 1) if user_total > 0 else 0
                
                # Last contact date for this user
                last_contact = user_query.with_entities(db.func.max(Lead.last_contact_date)).scalar()
                
                # Leads by stage
                stage_breakdown = user_query.with_entities(
                    Lead.stage,
                    db.func.count(Lead.id)
                ).group_by(Lead.stage).all()
                
                if user_total > 0:  # Only include users with leads
                    user_stats.append({
                        'user': user,
                        'total': user_total,
                        'active': user_active,
                        'converted': user_converted,
                        'lost': user_lost,
                        'conversion_rate': user_conv_rate,
                        'last_contact': last_contact,
                        'stage_breakdown': stage_breakdown
                    })
            
            # Sort by conversion rate (descending)
            user_stats.sort(key=lambda x: x['conversion_rate'], reverse=True)

        return render_template(
            "reports.html",
            total_leads=total_leads,
            active=active,
            converted=converted_total,
            lost=lost,
            conversion_rate=conversion_rate,
            source_rows=source_rows,
            course_rows=course_rows,
            all_users=all_users,
            selected_user_id=user_id_filter if user_id_filter else None,
            date_from=date_from,
            date_to=date_to,
            user_stats=user_stats
        )

    @app.route("/reports/export-excel", strict_slashes=False)
    @login_required
    @admin_required
    def reports_export_excel():
        """Export all database tables to Excel workbook sheets."""
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Define styles for headers
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        def apply_header_style(ws, headers_list):
            """Apply header styling to first row"""
            for col_num, header in enumerate(headers_list, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        def apply_cell_style(ws, row_num, col_count):
            """Apply borders to cells"""
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.border = border

        # ============ LEADS SHEET ============
        ws_leads = wb.create_sheet("Leads")
        
        leads_headers = [
            "ID", "Name", "Phone", "WhatsApp", "Gender", "Age",
            "Education Status", "Stream", "Institute", "Career Goal", 
            "Interested Courses", "Lead Source", "Decision Maker",
            "Start Timeframe", "Lead Score", "Stage", "Status",
            "Last Contact", "Next Followup", "Followup Count",
            "Notes", "Lost Reason", "Assigned To", "Created At", "Updated At"
        ]
        apply_header_style(ws_leads, leads_headers)
        
        all_leads = Lead.query.all()
        row = 2
        for lead in all_leads:
            ws_leads.cell(row=row, column=1).value = lead.id
            ws_leads.cell(row=row, column=2).value = lead.name
            ws_leads.cell(row=row, column=3).value = lead.phone
            ws_leads.cell(row=row, column=4).value = lead.whatsapp
            ws_leads.cell(row=row, column=5).value = lead.gender
            ws_leads.cell(row=row, column=6).value = lead.age
            ws_leads.cell(row=row, column=7).value = lead.education_status
            ws_leads.cell(row=row, column=8).value = lead.stream
            ws_leads.cell(row=row, column=9).value = lead.institute_name
            ws_leads.cell(row=row, column=10).value = lead.career_goal
            ws_leads.cell(row=row, column=11).value = lead.interested_courses
            ws_leads.cell(row=row, column=12).value = lead.lead_source
            ws_leads.cell(row=row, column=13).value = lead.decision_maker
            ws_leads.cell(row=row, column=14).value = lead.start_timeframe
            ws_leads.cell(row=row, column=15).value = lead.lead_score
            ws_leads.cell(row=row, column=16).value = lead.stage
            ws_leads.cell(row=row, column=17).value = lead.status
            ws_leads.cell(row=row, column=18).value = lead.last_contact_date.strftime('%d-%b-%Y') if lead.last_contact_date else ""
            ws_leads.cell(row=row, column=19).value = lead.next_followup_date.strftime('%d-%b-%Y') if lead.next_followup_date else ""
            ws_leads.cell(row=row, column=20).value = lead.followup_count
            ws_leads.cell(row=row, column=21).value = lead.notes
            ws_leads.cell(row=row, column=22).value = lead.lost_reason
            assigned_user = User.query.get(lead.assigned_to_id) if lead.assigned_to_id else None
            ws_leads.cell(row=row, column=23).value = assigned_user.full_name if assigned_user else ""
            ws_leads.cell(row=row, column=24).value = lead.created_at.strftime('%d-%b-%Y %H:%M') if lead.created_at else ""
            ws_leads.cell(row=row, column=25).value = lead.updated_at.strftime('%d-%b-%Y %H:%M') if lead.updated_at else ""
            
            apply_cell_style(ws_leads, row, len(leads_headers))
            row += 1
        
        # Set column widths
        ws_leads.column_dimensions['A'].width = 8
        ws_leads.column_dimensions['B'].width = 20
        ws_leads.column_dimensions['C'].width = 15
        ws_leads.column_dimensions['D'].width = 15
        ws_leads.column_dimensions['E'].width = 12
        ws_leads.column_dimensions['F'].width = 8
        ws_leads.column_dimensions['G'].width = 15
        ws_leads.column_dimensions['H'].width = 12
        ws_leads.column_dimensions['I'].width = 15
        ws_leads.column_dimensions['J'].width = 12
        ws_leads.column_dimensions['K'].width = 15
        ws_leads.column_dimensions['L'].width = 15
        ws_leads.column_dimensions['M'].width = 12
        ws_leads.column_dimensions['N'].width = 15
        ws_leads.column_dimensions['O'].width = 10
        ws_leads.column_dimensions['P'].width = 15
        ws_leads.column_dimensions['Q'].width = 12
        ws_leads.column_dimensions['R'].width = 12
        ws_leads.column_dimensions['S'].width = 12
        ws_leads.column_dimensions['T'].width = 12
        ws_leads.column_dimensions['U'].width = 20
        ws_leads.column_dimensions['V'].width = 15
        ws_leads.column_dimensions['W'].width = 20
        ws_leads.column_dimensions['X'].width = 18
        ws_leads.column_dimensions['Y'].width = 18

        # ============ USERS SHEET ============
        ws_users = wb.create_sheet("Users")
        
        users_headers = ["ID", "Username", "Full Name", "Role", "Is Active", "Created At"]
        apply_header_style(ws_users, users_headers)
        
        all_users = User.query.all()
        row = 2
        for user in all_users:
            ws_users.cell(row=row, column=1).value = user.id
            ws_users.cell(row=row, column=2).value = user.username
            ws_users.cell(row=row, column=3).value = user.full_name
            ws_users.cell(row=row, column=4).value = user.role
            ws_users.cell(row=row, column=5).value = "Yes" if user.is_active else "No"
            ws_users.cell(row=row, column=6).value = user.created_at.strftime('%d-%b-%Y %H:%M') if user.created_at else ""
            
            apply_cell_style(ws_users, row, len(users_headers))
            row += 1
        
        # Set column widths
        ws_users.column_dimensions['A'].width = 8
        ws_users.column_dimensions['B'].width = 15
        ws_users.column_dimensions['C'].width = 20
        ws_users.column_dimensions['D'].width = 12
        ws_users.column_dimensions['E'].width = 12
        ws_users.column_dimensions['F'].width = 18

        # ============ FOLLOWUPS SHEET ============
        ws_followups = wb.create_sheet("FollowUps")
        
        followups_headers = ["ID", "Lead ID", "Lead Name", "User", "Method", "Outcome", "Note", "Next Followup Date", "Created At"]
        apply_header_style(ws_followups, followups_headers)
        
        all_followups = FollowUp.query.all()
        row = 2
        for fu in all_followups:
            ws_followups.cell(row=row, column=1).value = fu.id
            ws_followups.cell(row=row, column=2).value = fu.lead_id
            ws_followups.cell(row=row, column=3).value = fu.lead.name if fu.lead else ""
            ws_followups.cell(row=row, column=4).value = fu.user.full_name if fu.user else ""
            ws_followups.cell(row=row, column=5).value = fu.method
            ws_followups.cell(row=row, column=6).value = fu.outcome
            ws_followups.cell(row=row, column=7).value = fu.note
            ws_followups.cell(row=row, column=8).value = fu.next_followup_date.strftime('%d-%b-%Y') if fu.next_followup_date else ""
            ws_followups.cell(row=row, column=9).value = fu.created_at.strftime('%d-%b-%Y %H:%M') if fu.created_at else ""
            
            apply_cell_style(ws_followups, row, len(followups_headers))
            row += 1
        
        # Set column widths
        ws_followups.column_dimensions['A'].width = 8
        ws_followups.column_dimensions['B'].width = 10
        ws_followups.column_dimensions['C'].width = 20
        ws_followups.column_dimensions['D'].width = 15
        ws_followups.column_dimensions['E'].width = 12
        ws_followups.column_dimensions['F'].width = 15
        ws_followups.column_dimensions['G'].width = 25
        ws_followups.column_dimensions['H'].width = 15
        ws_followups.column_dimensions['I'].width = 18

        # ============ ACTIVITIES SHEET ============
        ws_activities = wb.create_sheet("Activities")
        
        activities_headers = ["ID", "User", "Lead ID", "Lead Name", "Action Type", "Description", "Field Changed", "Old Value", "New Value", "Created At"]
        apply_header_style(ws_activities, activities_headers)
        
        all_activities = Activity.query.all()
        row = 2
        for activity in all_activities:
            ws_activities.cell(row=row, column=1).value = activity.id
            ws_activities.cell(row=row, column=2).value = activity.user.full_name if activity.user else ""
            ws_activities.cell(row=row, column=3).value = activity.lead_id
            ws_activities.cell(row=row, column=4).value = activity.lead.name if activity.lead else ""
            ws_activities.cell(row=row, column=5).value = activity.action_type
            ws_activities.cell(row=row, column=6).value = activity.description
            ws_activities.cell(row=row, column=7).value = activity.field_changed
            ws_activities.cell(row=row, column=8).value = activity.old_value
            ws_activities.cell(row=row, column=9).value = activity.new_value
            ws_activities.cell(row=row, column=10).value = activity.created_at.strftime('%d-%b-%Y %H:%M') if activity.created_at else ""
            
            apply_cell_style(ws_activities, row, len(activities_headers))
            row += 1
        
        # Set column widths
        ws_activities.column_dimensions['A'].width = 8
        ws_activities.column_dimensions['B'].width = 15
        ws_activities.column_dimensions['C'].width = 10
        ws_activities.column_dimensions['D'].width = 20
        ws_activities.column_dimensions['E'].width = 18
        ws_activities.column_dimensions['F'].width = 25
        ws_activities.column_dimensions['G'].width = 15
        ws_activities.column_dimensions['H'].width = 15
        ws_activities.column_dimensions['I'].width = 15
        ws_activities.column_dimensions['J'].width = 18

        # Save to BytesIO buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Return file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"leads_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )


    # -----------------------
    # ACTIVITY LOG (audit trail)
    # -----------------------
    @app.route("/activity-log")
    @login_required
    def activity_log():
        # Users see their own activities; admins can see all or filter by user
        date_from = request.args.get("date_from", "").strip()
        date_to = request.args.get("date_to", "").strip()
        user_id_filter = request.args.get("user_id", "").strip()
        action_type_filter = request.args.get("action_type", "").strip()
        
        query = Activity.query
        
        # Counselors see only their own activities
        if current_user.role == "counselor":
            query = query.filter(Activity.user_id == current_user.id)
            all_users = [current_user]
        else:
            # Admin can filter by user
            if user_id_filter:
                try:
                    user_id_filter = int(user_id_filter)
                    query = query.filter(Activity.user_id == user_id_filter)
                except (ValueError, TypeError):
                    pass
            
            all_users = User.query.order_by(User.full_name).all()
        
        # Date range filter
        if date_from:
            try:
                from_date = datetime.strptime(date_from, "%Y-%m-%d").date()
                query = query.filter(db.func.date(Activity.created_at) >= from_date)
            except ValueError:
                pass
        
        if date_to:
            try:
                to_date = datetime.strptime(date_to, "%Y-%m-%d").date()
                query = query.filter(db.func.date(Activity.created_at) <= to_date)
            except ValueError:
                pass
        
        # Action type filter
        if action_type_filter:
            query = query.filter(Activity.action_type == action_type_filter)
        
        # Get all activities sorted by newest first
        activities = query.order_by(Activity.created_at.desc()).all()
        
        # Get unique action types for dropdown
        all_action_types = db.session.query(Activity.action_type).distinct().order_by(Activity.action_type).all()
        all_action_types = [a[0] for a in all_action_types if a[0]]
        
        return render_template(
            "activity_log.html",
            activities=activities,
            all_users=all_users,
            all_action_types=all_action_types,
            date_from=date_from,
            date_to=date_to,
            selected_user_id=user_id_filter if user_id_filter else None,
            selected_action_type=action_type_filter,
            is_admin=(current_user.role == "admin")
        )

    # -----------------------
    # USERS (admin only)
    # -----------------------
    @app.route("/users")
    @login_required
    @admin_required
    def users_list():
        users = User.query.order_by(User.created_at.desc()).all()
        return render_template("users.html", users=users)

    @app.route("/users/new", methods=["POST"])
    @login_required
    @admin_required
    def users_create():
        username = request.form.get("username", "").strip()
        full_name = request.form.get("full_name", "").strip() or None
        role = request.form.get("role", "counselor").strip()
        password = request.form.get("password", "").strip()

        if not username or not password:
            flash("Username and password are required.", "danger")
            return redirect(url_for("users_list"))

        if role not in ["admin", "counselor"]:
            role = "counselor"

        if User.query.filter_by(username=username).first():
            flash("Username already exists.", "danger")
            return redirect(url_for("users_list"))

        u = User(username=username, full_name=full_name, role=role, is_active=True)
        u.set_password(password)
        db.session.add(u)
        db.session.commit()
        flash("User created.", "success")
        return redirect(url_for("users_list"))

    @app.route("/users/<int:user_id>/toggle", methods=["POST"])
    @login_required
    @admin_required
    def users_toggle(user_id):
        u = User.query.get_or_404(user_id)
        if u.id == current_user.id:
            flash("You cannot disable your own account.", "warning")
            return redirect(url_for("users_list"))

        u.is_active = not u.is_active
        db.session.commit()
        flash("User status updated.", "success")
        return redirect(url_for("users_list"))

    @app.route("/users/<int:user_id>/reset_password", methods=["POST"])
    @login_required
    @admin_required
    def users_reset_password(user_id):
        u = User.query.get_or_404(user_id)
        password = request.form.get("new_password", "").strip()
        if not password:
            flash("Password cannot be empty.", "danger")
            return redirect(url_for("users_list"))
        u.set_password(password)
        db.session.commit()
        flash("Password reset successful.", "success")
        return redirect(url_for("users_list"))

    return app


def seed_default_users():
    """Create default users once, if none exist."""
    if User.query.count() > 0:
        return

    # Default Admin
    admin = User(username="naveen", full_name="Naveen", role="admin", is_active=True)
    admin.set_password("admin123")

    # Default Counselor
    counselor = User(username="chaithra", full_name="Chaithra", role="counselor", is_active=True)
    counselor.set_password("chaithra123")

    db.session.add(admin)
    db.session.add(counselor)
    db.session.commit()


app = create_app()

if __name__ == "__main__":
    app.run(debug=True)
